"""
Excel 파일 생성 모듈
- 자동차 등록증 OCR 결과를 Excel로 저장
- 스타일링 (헤더, 테두리, 열 너비)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# 헤더 정의
HEADERS = [
    'No',
    '파일명',
    '자동차 등록번호',
    '차종',
    '차대번호',
    '소유자',
    '생년월일(법인등록번호)'
]

# 열 너비 (대략적)
COLUMN_WIDTHS = [6, 25, 20, 15, 25, 20, 25]


def create_excel(data: list[dict], output_path: str):
    """
    OCR 결과 데이터를 Excel 파일로 저장합니다.
    
    Args:
        data: OCR 결과 리스트. 각 항목은 {'filename': str, 'fields': dict} 형태
        output_path: 저장할 Excel 파일 경로 (.xlsx)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "자동차 등록증 정보"
    
    # --- 스타일 정의 ---
    header_font = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='맑은 고딕', size=10)
    cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    even_row_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    
    # --- 헤더 작성 ---
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # --- 데이터 작성 ---
    for row_idx, item in enumerate(data, 2):
        fields = item['fields']
        row_data = [
            row_idx - 1,                              # No
            item['filename'],                          # 파일명
            fields.get('자동차 등록번호', ''),          # 등록번호
            fields.get('차종', ''),                    # 차종
            fields.get('차대번호', ''),                # 차대번호
            fields.get('소유자', ''),                  # 소유자
            fields.get('생년월일(법인등록번호)', '')     # 생년월일
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            
            # No, 등록번호, 차종은 가운데 정렬, 나머지는 왼쪽
            if col_idx in (1, 3, 4):
                cell.alignment = cell_alignment
            else:
                cell.alignment = cell_alignment_left
            
            # 짝수 행 배경색
            if row_idx % 2 == 0:
                cell.fill = even_row_fill
    
    # --- 열 너비 설정 ---
    for col_idx, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # --- 행 높이 설정 ---
    ws.row_dimensions[1].height = 30  # 헤더
    for row_idx in range(2, len(data) + 2):
        ws.row_dimensions[row_idx].height = 25
    
    # --- 자동 필터 ---
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(data) + 1}"
    
    # --- 틀 고정 (헤더 행) ---
    ws.freeze_panes = 'A2'
    
    # 저장
    wb.save(output_path)
    wb.close()
